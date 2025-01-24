#!/usr/bin/env python

'''
This script is used to convert the Excel file with the keywords to a tab-delimited keywords file.
The script has two required arguments. ::

    Required:
    
    -e : The path to the XLSX file
    -o : The path to the output file
    
    Usage:
    
    python3 Excel2Keywords.py -e ../example/demo_keywords.xlsx -o ../YOUR_FOLDER/keywords_file.txt
    
'''

import argparse
import pandas as pd
import openpyxl
import json
from tqdm import tqdm

def process_excel(excel_file: str) -> pd.DataFrame:
    # Make docstring with rst syntax
    '''
    This function reads the Excel file and returns the data in the form of a DataFrame.\n
    \n
    Parameters:\n
    - excel_file: The path to the Excel file\n
    \n
    Returns:\n
    - df: A DataFrame with the columns (Category, Terms)
    '''
    
    # Read the Excel file sheetnames
    sheetnames = pd.ExcelFile(excel_file).sheet_names
    
    # Extract the relevant sheets (Human anatomy, Models, Omics, Diseases)
    # Regex on the sheet names should be used to extract the columns (since version numbers are also present)
    relevant_sheets = [sheet for sheet in sheetnames if "HUMAN ANATOMY" in sheet.upper() or "MODELS" in sheet.upper() or "OMICS" in sheet.upper() or "DISEASES" in sheet.upper()]
    
    # Read the relevant sheets
    df_dict = pd.read_excel(excel_file, sheet_name=relevant_sheets)
    
    # Convert to one DataFrame with columns (Category, Terms)
    df = pd.DataFrame(columns=["Category", "Terms"])
    
    # Concatenate the DataFrames
    for sheet in relevant_sheets:
        df_temp = df_dict[sheet]
        df_temp.columns = ["Category", "Terms"]
        df = pd.concat([df, df_temp], axis=0)
    
    return df

def clean_terms(df: pd.DataFrame) -> pd.DataFrame:
    # Make docstring with rst syntax
    """
    This function cleans the terms in the DataFrame.\n
    It removes, leading/trailing whitespaces, cleans hyphens, quotation marks, and removes duplicates.\n
    \n
    Parameters:\n
    - df: A pandas DataFrame\n
    \n
    Returns:\n
    - df_clean: A cleaned DataFrame
    """
    
    # Remove leading/trailing whitespaces
    df["Terms"] = df["Terms"].str.strip()
    
    # Remove hyphens
    df["Terms"] = df["Terms"].str.replace("-", " ")
    
    # Remove quotation marks
    df["Terms"] = df["Terms"].str.replace('"', "")
    
    # Remove duplicates
    df_clean = df.drop_duplicates()
    
    return df_clean

if __name__ == "__main__":
    
    # Create a parser object and add arguments
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("-e", dest="excel_file", required=True, help="Provide the path to the Excel file")
    parser.add_argument("-o", dest="output_file", required=True, help="Provide the path to the Output keyword file")

    # Read arguments from the command line
    args=parser.parse_args()
    
    # Read the Excel file
    df = process_excel(args.excel_file)
    
    # Clean the terms
    df_clean = clean_terms(df)
    
    # Save the DataFrame to a tab-delimited txt file without the column names
    df_clean.to_csv(args.output_file, sep="\t", index=False, header=False)
    
    
    

    