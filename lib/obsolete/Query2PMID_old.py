#!/usr/bin/env python

'''
This script is used to convert the queries from an Excel file to a list of PMIDs. It uses the NCBI API to query PubMed.

The script has three required arguments. ::

    Required:
    
    -q : The path to the queries .xlsx file (FOUND IN THE VALIDATION FOLDER)
    -o : The path to the output directory
    -e : The email address for the NCBI API
    
    Usage:
    
    python3 Query2PMID.py -q ../example/queries/queries.xlsx -o ../example/queries/ -e EMAIL@EMAIL.com
    
'''

# Import the required libraries
from Bio import Entrez
from Bio import Medline
from tqdm import tqdm
from openpyxl import load_workbook
import argparse
import time

def run_query(query: str) -> list:
    """
    Function to run a query on PubMed\n
    \n
    Parameters:\n
    - query: The query to run\n
    \n
    Returns:\n
    - The list of PMIDs
    """
    # Run the query
    handle = Entrez.esearch(db="pubmed", retmax=2500, term=query)
    record = Entrez.read(handle)
    handle.close()
    
    # Get the PMIDs
    pmids = record['IdList']
    return pmids
    
if __name__ == "__main__":
    
    # Create a parser object and add arguments
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("-q", dest="query_file",       required=True, help="An Excel file with queries")
    parser.add_argument("-o", dest="output_directory", required=True, help="Provide the path to the Output JSON file")
    parser.add_argument("-e", dest="entrez_email",     required=True, help="Provide the email address for the NCBI API")
    
    # Read arguments from the command line
    args=parser.parse_args()
    Entrez.email = args.entrez_email
    
    # Load the Excel file
    workbook = load_workbook(args.query_file)
    sheet = workbook.active

    query_dict = {}

    # Iterate through rows starting from the second row (to skip the header if there is one)
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key = row[0]  # First column as key
        value = row[1]  # Second column as value
        query_dict[key] = value

    # Display the resulting dictionary

    # Set a time limit
    TIME_LIMIT =  '("2024/01/01"[Date - Publication] : "3000"[Date - Publication])'

    # Get model query 
    MODEL_QUERY = query_dict['models']
    
    # Create the queries for human only and human and animal studies
    queries = {
        "human_only": query_dict['human'],
        "human_animal": query_dict['human_animal']
    }

    # Cycle through the disease areas and execute queries
    disease_areas = ['atmps', 'autoimmun','breastcanc','cardiovasc','gastrointestin','immonco','infectdis','metadisord','neurodeg','respdis']
    # disease_areas = ['atmps'] # FOR TESTING
    
    pmids = []
    
    # Cycle through the queries and disease areas and execute the queries
    for query_type in queries:
        for disease_area in tqdm(disease_areas):
            tqdm.write(f"Querying {query_type} for {disease_area}")
            full_query   = "(" + query_dict[disease_area] + ") AND (" + queries[query_type] + ") AND (" + MODEL_QUERY + ") AND (" + TIME_LIMIT + ")"
            
            new_pmids = run_query(query = full_query)
            to_save = [(disease_area, query_type, pmid) for pmid in new_pmids]

            pmids.extend(to_save)
            time.sleep(2)
        
    # Create unique list of pmids for further processing
    pmid_list = list(set([str(pmid_tuple[2]) for pmid_tuple in pmids]))
    OUT = open(args.output_directory +"predict_pmids.txt","w")
    OUT.write("\n".join(pmid_list))
    OUT.close()
        
    # Keep mapping of pmids to disease area and query type
    OUT = open(args.output_directory +"pmid_mapping.txt","w")
    OUT.write("\n".join("\t".join(map(str, pmid_tuple)) for pmid_tuple in pmids))
    OUT.close()