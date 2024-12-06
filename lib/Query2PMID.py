#!/usr/bin/env python

'''

This script executes a set of queries that are supplied in an Excel file. For each query, a result file is created in the output directory specified by the -o option. The result file is named after the query_id that is specified in the first column of the Excel file.

The script has three required arguments. ::

    Required:
    
    -q : The path to the Excel file with the queries
    -o : The path to the output directory    
    -e : The email address for the NCBI API
    
    Usage:
    
    ./Query2PMID.py -q ../example/example_query.xlsx -o /tmp/  -e youremail@email.com
    
'''

# Import the required libraries
from Bio import Entrez
from Bio import Medline

import sys
from openpyxl import load_workbook
import argparse
import time

from tqdm import tqdm

from datetime import datetime
from datetime import timedelta


    
if __name__ == "__main__":
    
    # Create a parser object and add arguments
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("-q", dest="query_file", required=True, help="An Excel file with queries")
    parser.add_argument("-o", dest="output_directory", required=True, help="Provide the path to the Output JSON file")
    parser.add_argument("-e", dest="entrez_email", required=True, help="Provide the email address for the NCBI API")
    parser.add_argument("-r", dest="retmax", required=True, help="Provide the maximum number of records to retrieve")

    # Read arguments from the command line
    args=parser.parse_args()

    # Load the Excel file
    workbook = load_workbook(args.query_file)

   
    # Select the active sheet or specify a sheet by name
    sheet = workbook.active

    # Iterate through rows in the sheet
    queries = {}
    for row in sheet.iter_rows(min_row=2,values_only=True):
        full_query = row[1]
        # Add the query to the dictionary with row[0] as the key (disease area)
        queries[row[0]] = full_query


       
    Entrez.email = args.entrez_email

    for q in queries.keys():
        print("Running query " + queries[q])
        OUT = open(args.output_directory+"/"+q+"_predict_pmids.txt","w")
        
        retmax = int(args.retmax)
        
        # If retmax is larger than 10000, we need to split the query with retstart increments of 10000
        if retmax > 10000:
            print("Retmax is larger than 10000")
            
            # Get today as the end date
            end_date = datetime.now().strftime("%Y/%m/%d")
            
            # By default, start date is 7 days before the end date
            start_date = (datetime.strptime(end_date, "%Y/%m/%d") - timedelta(days=7)).strftime("%Y/%m/%d")
            all_pmids = []
            
            while len(list(set(all_pmids))) < retmax:                
                query = f"{queries[q]} AND ({start_date}[PDAT] : {end_date}[PDAT])"
                handle = Entrez.esearch(
                    db="pubmed",
                    term=query,
                    retmax=10000,
                    usehistory="y"
                )
                record = Entrez.read(handle)
                handle.close()
                pmids = record["IdList"]
                
                
                
                if len(pmids) == 9999:
                    print(f"Retrieved {len(pmids)} records from {start_date} to {end_date}")
                    print("Splitting date range in half")
                    start_date = (datetime.strptime(end_date, "%Y/%m/%d") - timedelta(days=2)).strftime("%Y/%m/%d")
                else:
                    print(f"Retrieved {len(pmids)} records from {start_date} to {end_date}")
                    all_pmids.extend(pmids)
                    # Create new start and end dates 1 week before the previous start date and end date
                    end_date = start_date
                    start_date = (datetime.strptime(end_date, "%Y/%m/%d") - timedelta(days=7)).strftime("%Y/%m/%d")
                time.sleep(2)
                
            all_pmids = list(set(all_pmids)) 
            OUT.write("\n".join(str(x) for x in all_pmids))         
            time.sleep(2)
        
        else:
            handle = Entrez.esearch(db="pubmed", retmax=args.retmax, term=queries[q])
            record = Entrez.read(handle)
            handle.close()
            OUT.write("\n".join(str(x) for x in (record['IdList'])))
            OUT.close()
            time.sleep(2)
            
        # Check how many results are in the output file
        with open(args.output_directory+"/"+q+"_predict_pmids.txt") as f:
            lines = f.readlines()
            print("Number of PMIDs found: " + str(len(lines)))
        
